import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import subprocess
import json
import os
import xml.etree.ElementTree as ET
from pathlib import Path

# DEBUG MODE - Set to True to only process first x objects for faster testing
DEBUG_MODE = False
MAX_OBJECTS_DEBUG = 5

# Path to your SFDX source - adjust if needed
SFDX_SOURCE_PATH = 'uksa-sm/main/default'

def get_all_objects():
    """Get list of all custom and standard objects"""
    print("Fetching all objects...")
    query = """
    SELECT QualifiedApiName, Label, DeveloperName 
    FROM EntityDefinition 
    WHERE IsCustomizable = true
    ORDER BY QualifiedApiName
    """
    
    result = subprocess.run(
        ['sf', 'data', 'query', '--query', query, '--json'],
        capture_output=True,
        text=True
    )
    
    if result.returncode != 0:
        print(f"Error fetching objects: {result.stderr}")
        return []
    
    data = json.loads(result.stdout)
    return data.get('result', {}).get('records', [])

def get_global_value_sets_from_source():
    """Get global value sets from SFDX source metadata files"""
    print("\nFetching global value sets from source...")
    
    value_sets_path = Path(SFDX_SOURCE_PATH) / 'globalValueSets'
    
    if not value_sets_path.exists():
        print(f"  Warning: Global value sets directory not found at {value_sets_path}")
        return []
    
    rows = []
    
    for vs_file in value_sets_path.glob('*.globalValueSet-meta.xml'):
        try:
            tree = ET.parse(vs_file)
            root = tree.getroot()
            
            # Remove namespace if present
            ns = {'sf': 'http://soap.sforce.com/2006/04/metadata'}
            
            # Get value set name from filename
            vs_name = vs_file.stem.replace('.globalValueSet-meta', '')
            
            # Get master label
            master_label = root.find('.//sf:masterLabel', ns)
            if master_label is None:
                master_label = root.find('.//masterLabel')
            
            # Get description
            description = root.find('.//sf:description', ns)
            if description is None:
                description = root.find('.//description')
            
            # Get all custom values
            values = []
            for value_elem in root.findall('.//sf:customValue', ns):
                full_name = value_elem.find('sf:fullName', ns)
                is_active = value_elem.find('sf:isActive', ns)
                
                if full_name is not None and (is_active is None or is_active.text == 'true'):
                    values.append(full_name.text)
            
            # Try without namespace if no values found
            if not values:
                for value_elem in root.findall('.//customValue'):
                    full_name = value_elem.find('fullName')
                    is_active = value_elem.find('isActive')
                    
                    if full_name is not None and (is_active is None or is_active.text == 'true'):
                        values.append(full_name.text)
            
            rows.append({
                'Value Set Name': vs_name,
                'Label': master_label.text if master_label is not None else vs_name,
                'Description': description.text if description is not None else '',
                'Number of Values': len(values),
                'Values': '; '.join(values),
                'Last Technical Update': datetime.now().strftime('%Y-%m-%d')
            })
            
        except Exception as e:
            if DEBUG_MODE:
                print(f"  Error reading {vs_file.name}: {str(e)[:100]}")
    
    print(f"  Found {len(rows)} global value sets in source")
    return rows

def get_global_value_sets():
    """Get global value sets - wrapper that uses source metadata"""
    return get_global_value_sets_from_source()

def get_fields_for_object(object_name):
    """Get all fields for a specific object with detailed metadata"""
    print(f"  Fetching fields for {object_name}...")
    query = f"""
    SELECT 
        QualifiedApiName,
        Label,
        Description,
        NamespacePrefix,
        DataType,
        Length,
        IsIndexed,
        IsNillable,
        Precision,
        Scale,
        IsCalculated,
        IsCompound,
        ExtraTypeInfo
    FROM FieldDefinition 
    WHERE EntityDefinition.QualifiedApiName='{object_name}'
    ORDER BY QualifiedApiName
    """
    
    result = subprocess.run(
        ['sf', 'data', 'query', '--query', query, '--json'],
        capture_output=True,
        text=True
    )
    
    if result.returncode != 0:
        print(f"    Warning: Could not fetch fields for {object_name}")
        if result.stderr:
            print(f"    Error details: {result.stderr[:200]}")
        return []
    
    data = json.loads(result.stdout)
    return data.get('result', {}).get('records', [])

def get_picklist_values_from_source(object_name, field_name):
    """Get picklist values from SFDX source metadata files"""
    try:
        # Look for the field file in the source
        field_path = Path(SFDX_SOURCE_PATH) / 'objects' / object_name / 'fields' / f'{field_name}.field-meta.xml'
        
        if not field_path.exists():
            return '[Field metadata not found in source]'
        
        # Parse the XML
        tree = ET.parse(field_path)
        root = tree.getroot()
        
        # Remove namespace if present
        ns = {'sf': 'http://soap.sforce.com/2006/04/metadata'}
        
        # Check for value set reference (global value set)
        value_set = root.find('.//sf:valueSet/sf:valueSetName', ns)
        if value_set is None:
            value_set = root.find('.//valueSet/valueSetName')
        
        if value_set is not None and value_set.text:
            return f"[Global Value Set: {value_set.text}]"
        
        # Get individual picklist values
        values = []
        for value_elem in root.findall('.//sf:valueSetDefinition/sf:value', ns):
            full_name = value_elem.find('sf:fullName', ns)
            is_active = value_elem.find('sf:isActive', ns)
            
            if full_name is not None and (is_active is None or is_active.text == 'true'):
                values.append(full_name.text)
        
        # If no values found with namespace, try without
        if not values:
            for value_elem in root.findall('.//valueSetDefinition/value'):
                full_name = value_elem.find('fullName')
                is_active = value_elem.find('isActive')
                
                if full_name is not None and (is_active is None or is_active.text == 'true'):
                    values.append(full_name.text)
        
        if values:
            return '; '.join(values)
        
        return '[No values found in source metadata]'
        
    except Exception as e:
        if DEBUG_MODE:
            print(f"      Exception reading source for {field_name}: {str(e)[:100]}")
        return '[Error reading source metadata]'

def get_picklist_values(object_name, field_name):
    """Get picklist values - wrapper function that tries source first"""
    return get_picklist_values_from_source(object_name, field_name)

def parse_salesforce_metadata():
    """Query Salesforce org for comprehensive metadata"""
    rows = []
    
    # Get all objects
    objects = get_all_objects()
    
    if not objects:
        print("No objects found. Make sure you're authenticated to a Salesforce org.")
        print("Run: sf org login web")
        return []
    
    # Limit objects in debug mode
    if DEBUG_MODE:
        objects = objects[:MAX_OBJECTS_DEBUG]
        print(f"\n*** DEBUG MODE: Processing only first {MAX_OBJECTS_DEBUG} objects ***\n")
    
    print(f"\nFound {len(objects)} objects. Fetching field details...")
    
    # For each object, get all fields
    for obj in objects:
        object_api_name = obj.get('QualifiedApiName', '')
        object_label = obj.get('Label', '')
        
        fields = get_fields_for_object(object_api_name)
        
        for field in fields:
            field_api_name = field.get('QualifiedApiName', '')
            data_type = field.get('DataType', '')
            
            # Determine if field is required (not nillable)
            is_required = 'Yes' if not field.get('IsNillable', True) else 'No'
            
            # Get length information
            length = field.get('Length', '')
            if field.get('Precision') and field.get('Scale'):
                length = f"{field.get('Precision')},{field.get('Scale')}"
            
            # Determine if custom field (must end with __c)
            is_custom = field_api_name.endswith('__c')
            field_type = 'Custom Field' if is_custom else 'Standard Field'
            
            # Get extra type info for special fields
            extra_info = field.get('ExtraTypeInfo', '')
            if extra_info:
                data_type = f"{data_type} ({extra_info})"
            
            # Get picklist values only for custom fields without namespace
            picklist_values = ''
            if data_type in ['Picklist', 'MultiselectPicklist']:
                namespace = field.get('NamespacePrefix', '')
                
                # Only fetch for custom fields (ends with __c) without a namespace
                if is_custom and not namespace:
                    print(f"    Fetching picklist values for {field_api_name}...")
                    picklist_values = get_picklist_values(object_api_name, field_api_name)
                elif namespace:
                    picklist_values = '[Managed package field - see package documentation]'
                else:
                    picklist_values = '[Standard field - see Salesforce documentation]'
            
            rows.append({
                'Object API Name': object_api_name,
                'Object Label': object_label,
                'Field API Name': field_api_name,
                'Field Label': field.get('Label', ''),
                'Namespace': field.get('NamespacePrefix', ''),
                'Type': field_type,
                'Data Type': data_type,
                'Required': is_required,
                'Length': str(length) if length else '',
                'Indexed': 'Yes' if field.get('IsIndexed', False) else 'No',
                'Calculated': 'Yes' if field.get('IsCalculated', False) else 'No',
                'Picklist Values': picklist_values,
                'Description': field.get('Description', ''),
                'Business Purpose': '',
                'Data Owner': '',
                'Related Processes': '',
                'Last Technical Update': datetime.now().strftime('%Y-%m-%d'),
                'Last Business Review': ''
            })
    
    return rows

def format_excel(filename):
    """Apply formatting to the Excel file"""
    wb = load_workbook(filename)
    
    # Format Fields sheet
    ws_fields = wb['Fields']
    
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws_fields[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Set column widths for Fields sheet
    column_widths = {
        'A': 20,  # Object API Name
        'B': 20,  # Object Label
        'C': 25,  # Field API Name
        'D': 20,  # Field Label
        'E': 12,  # Namespace
        'F': 15,  # Type
        'G': 20,  # Data Type
        'H': 10,  # Required
        'I': 10,  # Length
        'J': 10,  # Indexed
        'K': 12,  # Calculated
        'L': 50,  # Picklist Values
        'M': 40,  # Description
        'N': 40,  # Business Purpose
        'O': 20,  # Data Owner
        'P': 30,  # Related Processes
        'Q': 18,  # Last Technical Update
        'R': 18   # Last Business Review
    }
    
    for col, width in column_widths.items():
        ws_fields.column_dimensions[col].width = width
    
    # Freeze header row and first two columns
    ws_fields.freeze_panes = 'C2'
    
    # Enable autofilter
    ws_fields.auto_filter.ref = ws_fields.dimensions
    
    # Format Value Sets sheet if it exists
    if 'Global Value Sets' in wb.sheetnames:
        ws_vs = wb['Global Value Sets']
        
        for cell in ws_vs[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Set column widths for Value Sets sheet
        vs_column_widths = {
            'A': 30,  # Value Set Name
            'B': 30,  # Label
            'C': 40,  # Description
            'D': 15,  # Number of Values
            'E': 60,  # Values
            'F': 18   # Last Technical Update
        }
        
        for col, width in vs_column_widths.items():
            ws_vs.column_dimensions[col].width = width
        
        ws_vs.freeze_panes = 'A2'
        ws_vs.auto_filter.ref = ws_vs.dimensions
    
    wb.save(filename)

def main():
    print("Salesforce Org Data Dictionary Generator")
    print("=" * 50)
    
    if DEBUG_MODE:
        print(f"\n*** DEBUG MODE ENABLED - Processing only {MAX_OBJECTS_DEBUG} objects ***")
        print("*** Set DEBUG_MODE = False in the script for full run ***\n")
    
    print("\nThis will query your authenticated Salesforce org")
    print("Make sure you're logged in: sf org login web")
    print()
    
    # Parse metadata from Salesforce
    rows = parse_salesforce_metadata()
    
    if not rows:
        print("\nError: No metadata found. Please check:")
        print("  1. You're authenticated to a Salesforce org")
        print("  2. The Salesforce CLI is installed and working")
        print("  3. Run: sf org list to see available orgs")
        return
    
    # Get global value sets
    value_set_rows = get_global_value_sets()
    
    if DEBUG_MODE:
        print(f"\n*** DEBUG: Found {len(value_set_rows)} global value sets ***")
    
    # Create DataFrames
    df_fields = pd.DataFrame(rows)
    df_fields = df_fields.sort_values(['Object API Name', 'Field API Name'])
    
    # Create Excel file with multiple sheets
    output_file = 'salesforce_data_dictionary.xlsx'
    print(f"\nCreating Excel file: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_fields.to_excel(writer, index=False, sheet_name='Fields')
        
        if value_set_rows:
            df_value_sets = pd.DataFrame(value_set_rows)
            df_value_sets = df_value_sets.sort_values('Value Set Name')
            df_value_sets.to_excel(writer, index=False, sheet_name='Global Value Sets')
    
    # Apply formatting
    print("Applying formatting...")
    format_excel(output_file)
    
    print(f"\n✓ Success! Created {output_file}")
    print(f"\nFields Sheet:")
    print(f"  - Total fields: {len(rows)}")
    print(f"  - Unique objects: {df_fields['Object API Name'].nunique()}")
    print(f"  - Custom fields: {len(df_fields[df_fields['Type'] == 'Custom Field'])}")
    print(f"  - Standard fields: {len(df_fields[df_fields['Type'] == 'Standard Field'])}")
    
    if value_set_rows:
        print(f"\nGlobal Value Sets Sheet:")
        print(f"  - Total value sets: {len(value_set_rows)}")
    
    print("\nNext steps:")
    print("  1. Review the generated file")
    print("  2. Fill in Business Purpose, Data Owner, and Related Processes columns")
    print("  3. Update or enhance Description fields where needed")
    print("  4. Update Last Business Review dates as you complete reviews")

if __name__ == "__main__":
    main()